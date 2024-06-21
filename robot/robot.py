import math
from datetime import datetime
import openpyxl
from browser.chrome import ChromeBrowser
from files_and_folders.folders import Folder
from robot_manager.base import Bot
import pandas as pd
from robot.app import App
from .constants import *
from .flow import *
from .exceptions import *

class Robot(Bot):
    """
    Este robot se encarga de descargar los trámites pendientes en la web de la agencia tributaria de Vizcaya,
    Guardarlos en la ruta corresponmdiente y enviarselos a los clientes
    """

    def __init__(self, **kwargs):
        super().__init__(**kwargs, disabled=False)
        self.transaction_number = None
        self.data = None
        self.app = None
        self.browser = None
        self.end_date = None
        self.start_date = None
        self.folder = None
        self.tempFolder = None
        self.nif = None

    @RobotFlow(Nodes.StartNode, children="get_client_data")
    def start(self):
        """
        En este método se inicializa el robot, se definen las variables globales como rutas de descargas,
        datos de entrada y otras variables que se utilizarán durante el proceso y se inicia sesión en la página web de la agencia tributaria de Vizcaya.

        Excepciones:
            1. Excepciones de Sistema:
                1.1 Cualquier excepción -> next_action: "retry"
        """

        try:
            self.tempFolder = TEMP_FOLDER
            self.workbook_path = INPUT_FILE
            self.folder = Folder(self.tempFolder)
            self.folder.empty(allow_root=True)
            try:
                self.start_date = datetime.strptime(self.parameters.get('date-from'), '%Y-%m-%d').strftime('%d/%m/%Y')
                self.end_date = datetime.strptime(self.parameters.get('date-to'),  '%Y-%m-%d').strftime('%d/%m/%Y')
            except:
                self.start_date = datetime.now().strftime('%d/%m/%Y')
                self.end_date = datetime.now().strftime('%d/%m/%Y')
            self.log.trace(f"Se van a obtener los impuestos desde {self.start_date} hasta {self.end_date}")
            self.browser = ChromeBrowser(undetectable=True)
            self.browser.options.page_load_strategy = "normal"
            self.browser.options.add_experimental_option("prefs", {
                "download.prompt_for_download": False,  # Desactiva el diálogo de confirmación de descarga
                "download.directory_upgrade": True,
                "plugins.always_open_pdf_externally": True,
                "download.default_directory": f"{self.tempFolder}",
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "safebrowsing_for_trusted_sources_enabled": False,
                "safebrowsing.enabled": False
            })
            self.app = App(self.browser)
            self.app.login()
            self.data = pd.read_excel(self.workbook_path)
            self.transaction_number = 0
            self.wb = openpyxl.load_workbook(self.workbook_path)
        except Exception as e:
            self.browser.close()
            self.log.system_exception(e)
            raise SystemException(self, message=e, next_action="retry")

    @RobotFlow(Nodes.ConditionNode, children={True: "get_client_documents", False: "end"},
               condition=Conditions.has_data)
    def get_client_data(self, *args):
        """
        A partir de la variable data (self.data) creada en el método anterior,
        se gestionan las transacciones una a una para procesarla por los siguiente métodos:
            1. Si existen transacciones pendientes de procesar: se llama al método "get_cliente_documents",
            2. En caso contrario termina el proceso método "end"
        """
        if self.data.empty:
            return []
        return self.data.iloc[0]



    @RobotFlow(Nodes.OperationNode, children="process_documents")
    def get_client_documents(self, *args):
        """
        En este método se busca al cliente y se filtran los expediente del mes correspondiente.
        Recibe como argumento una la información del cliente.
        Devuelve una lista con los trámites disponibles para descargar.

        Exceciones:
            1. Excepciones de Negocio:
                1.1 No se encuentra al usuario -> next_action: "set_transaction_status"
                1.2 No hay trámites para el usuario -> next_action: "set_transaction_status"
            2. Excepciones de Sistema:
                Cualquier excepcion no controlada -> next_action: "retry"

        """
        try:
            self.cod_cliente = args[0].iloc[0]
            self.nif = args[0].iloc[1]
            self.name = args[0].iloc[2]
            self.downloaded_documents = []
            self.log.debug(f"se va a procesar el {self.nif}")
            self.transaction_number = self.transaction_number + 1
            self.current_page =1

            if self.app.find_client(self.nif) is False:
                message = f"No se encuentra el usuario {self.nif}"
                raise BusinessException(self, message=message, next_action="set_transaction_status")
            time.sleep(10)
            tramites = self.app.filter_data(self.start_date, self.end_date)
            if len(tramites) == 0:
                message = "No hay tramites para el usuario"
                raise BusinessException(self, message=message, next_action="set_transaction_status")
            self.folder.empty()
            #Si
            if len(self.folder.file_list(".pdf")) > 0:
                self.folder.empty()
                time.sleep(30)
                if len(self.folder.file_list(".pdf")) > 0:
                    raise BusinessException(self, message="Error al borrar documentos de la carpeta", next_action="retry")
            return tramites

        except BusinessException as BE:
            self.log.business_exception(BE.message)
            raise BE
        except Exception as e:
            self.log.system_exception("Error al obtener los documentos del cliente: Reintentando")
            self.browser.get("https://ataria.ebizkaia.eus/es/mis-presentaciones")
            raise SystemException(self, message=e, next_action="retry")

    @RobotFlow(Nodes.ConditionNode, children={True: "download_document",False: "set_transaction_status"},
               condition=Conditions.has_data)
    def process_documents(self, *args):
        """
        Se evalúa si quedan trámites disponibles del cliente para descargar, en caso de que existan,
        Se procederá a llamar al nodo de "download_document.
        En caso contrario, se gestionará la transacción en el nodo "set_transaction_status"
        """
        if args:
            tramites = args[0]
            if len(tramites) > 0:
                if math.ceil((tramites[0]+1)/10) > self.current_page:
                    try:
                        self.browser.find_element(by="xpath",
                                              value="//*[@class='ui-paginator-next ui-state-default ui-corner-all']").click()
                        self.current_page += 1
                        time.sleep(10)
                    except Exception as e:
                        raise SystemException(self, message=e, next_action="retry")
            return tramites
        else:
            return []

    @RobotFlow(Nodes.OperationNode, children="process_documents")
    def download_document(self, *args):
        """
        Este método recibe como argumento el trámite que se tiene que procesar.
        Descarga el documento y lee la información de este.

        Excepciones:
            1. Negocio: No hay documentos para descargar -> next_action : "set_transaction_status"
            2. Sistema: No se ha descargadi el documento -> next_action: "rety"
        """
        try:
            tramites = args[0]
            tramite = tramites[0]

            self.app.descargar_documentos(tramite)
            if len(self.folder.file_list(".pdf")) == 0:
                if len(self.folder.file_list(".PDF")) == 0:
                    raise SystemException(self, message="No se puede descargar el documento", next_action="retry")
                else:
                    file = self.folder.file_list(".PDF")[0]
            else:
                file = self.folder.file_list(".pdf")[0]
            try:
                impuesto = self.app.save_file(file, self.name, self.nif, self.cod_cliente)


            except NameError as e:
                raise BusinessException(self, message=e, next_action="skip")
            except Exception as e:
                raise e

            self.log.trace(impuesto[2])
            page = self.wb['IMPUESTOS']
            page.append([self.nif, self.name, impuesto[0], impuesto[1], impuesto[2]])
            self.wb.save(self.workbook_path)
            self.wb.close()
            self.folder.empty()
            tramites.pop(0)

            archivo = impuesto[2]
            self.downloaded_documents.append(archivo)
            return tramites

        except BusinessException as BE:
            try:
                self.folder.empty(allow_root=True)
            except:
                raise SystemException(self, message=e, next_action="retry")
            self.log.business_exception(BE.message)
            tramites.pop(0)
            raise BE

        except Exception as e:
            self.log.system_exception(e)
            try:
                self.folder.empty(allow_root=True)
            except:
                raise SystemException(self, message=e, next_action="retry")
            raise SystemException(self, message=e, next_action="retry")

    @RobotFlow(Nodes.OperationNode, children="get_client_data")
    def set_transaction_status(self, *args):
        # Remove Processed Item
        if len(args[0]) > 0:
            result = args[0][0]
        else:
            result = "OK"
            self.log.trace(result)
            # Selecciona la hoja en la que quieres buscar y actualizar
        hoja = self.wb['CLIENTES']
        # Dato que estás buscando
        dato_a_buscar = self.nif
        # Busca el dato en la columna A y obtén la fila
        fila = None
        for fila_actual in range(1, hoja.max_row + 1):
            if hoja.cell(row=fila_actual, column=1).value == dato_a_buscar:
                fila = fila_actual
                break
        # Si se encuentra el dato, actualiza la celda correspondiente en la columna D
        if fila is not None:
            hoja.cell(row=fila, column=7).value = result
            hoja.cell(row=fila, column=8).value = datetime.today().strftime("%d-%m-%y")
        # Guarda los cambios en el archivo
        self.wb.save(INPUT_FILE)
        self.wb.close()

        self.data = self.data.drop(0)
        self.data.reset_index(drop=True, inplace=True)

    @RobotFlow(node=Nodes.EndNode)
    def end(self, *args):
        """
        Este método cierra el Navegador web.
        """
        self.browser.close()
        self.log.trace(f"end")
